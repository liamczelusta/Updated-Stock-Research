"""Fixed-layout Excel parser for the stock research workbook template."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from stock_research.dataclasses import (
    AdditionalMetricValue,
    CompanyInfo,
    ParsedWorkbook,
    ParserWarning,
    QuarterRecord,
    StaticCell,
)


class WorkbookValidationError(ValueError):
    """Raised when the workbook does not match the expected fixed layout."""


@dataclass(frozen=True)
class ColumnSpec:
    """Expected fixed column metadata."""

    column: int
    name: str
    required_header_tokens: tuple[str, ...] = ()
    formula_required: bool = False


DATA_SHEET_NAME = "Sheet1"
EXPECTED_SHEETS = ("Sheet1", "Sheet2", "Sheet3")
HEADER_ROW_CANDIDATES = (3, 4)
DATA_START_ROW = 6
FOOTER_PREFIX = "the report herein"

CORE_COLUMNS = {
    "quarter": ColumnSpec(1, "Quarter", ("quarter",)),
    "status": ColumnSpec(2, "B/A", ()),
    "revenue_estimate": ColumnSpec(3, "CR / mil.", ("cr",)),
    "revenue_actual": ColumnSpec(4, "Actual revenue", ("actual", "rev")),
    "rolling_revenue": ColumnSpec(5, "Rolling 4 qtrs.", ("rolling",)),
    "rolling_revenue_yoy": ColumnSpec(6, "Rolling revenue YOY", ("yoy",)),
    "revenue_surprise": ColumnSpec(7, "Diff. fr CR", ("diff",)),
    "revenue_qoq": ColumnSpec(8, "Revenue QoQ", ()),
    "revenue_yoy": ColumnSpec(9, "Revenue YoY", ("yoy",)),
    "eps_estimate": ColumnSpec(10, "CE", ("ce",)),
    "eps_actual": ColumnSpec(11, "Actual earnings", ("actual",)),
    "rolling_eps": ColumnSpec(12, "Rolling 4 qtrs. EPS", ("rolling",)),
    "rolling_eps_yoy": ColumnSpec(13, "Rolling EPS YOY", ("yoy",)),
    "eps_surprise": ColumnSpec(14, "Diff. fr CE", ("diff",)),
    "eps_qoq": ColumnSpec(15, "EPS QoQ", ()),
    "eps_yoy": ColumnSpec(16, "EPS YoY", ("yoy",)),
    "total_earnings": ColumnSpec(17, "Total earnings / shares header", ()),
    "price_before": ColumnSpec(18, "Price before", ("price", "before")),
    "price_after": ColumnSpec(19, "Price after", ("price", "after")),
    "price_reaction": ColumnSpec(20, "Price diff.", ("price", "diff")),
    "shares_or_notes": ColumnSpec(21, "Shares / notes", ()),
}

FORMULA_COLUMNS = {
    "rolling_revenue",
    "rolling_revenue_yoy",
    "revenue_surprise",
    "revenue_qoq",
    "revenue_yoy",
    "rolling_eps",
    "rolling_eps_yoy",
    "eps_surprise",
    "eps_qoq",
    "eps_yoy",
    "total_earnings",
    "price_reaction",
}

FORECAST_INPUT_COLUMNS = (
    CORE_COLUMNS["revenue_estimate"].column,
    CORE_COLUMNS["revenue_actual"].column,
    CORE_COLUMNS["eps_estimate"].column,
    CORE_COLUMNS["eps_actual"].column,
)


def parse_workbook(path: str | Path) -> ParsedWorkbook:
    """Parse one fixed-format stock workbook into typed dataclasses."""

    workbook_path = Path(path).expanduser()
    if not workbook_path.exists():
        raise WorkbookValidationError(f"Workbook not found: {workbook_path}")

    formula_wb = load_workbook(workbook_path, data_only=False, read_only=False)
    value_wb = load_workbook(workbook_path, data_only=True, read_only=False)

    _validate_sheets(formula_wb.sheetnames)
    formula_ws = formula_wb[DATA_SHEET_NAME]
    value_ws = value_wb[DATA_SHEET_NAME]

    header_row = _find_header_row(value_ws)
    _validate_core_headers(value_ws, header_row)

    data_end_row = _find_data_end_row(value_ws)
    if data_end_row < DATA_START_ROW:
        raise WorkbookValidationError(
            f"{DATA_SHEET_NAME} has no quarterly rows starting at row {DATA_START_ROW}."
        )

    warnings: list[ParserWarning] = []
    additional_headers = _read_additional_headers(value_ws, header_row)
    static_cells = _read_static_cells(value_ws, header_row)
    company = _read_company_info(
        workbook_path=workbook_path,
        sheet_names=tuple(formula_wb.sheetnames),
        header_row=header_row,
        data_end_row=data_end_row,
        static_cells=tuple(static_cells),
    )

    quarters: list[QuarterRecord] = []
    for row_number in range(DATA_START_ROW, data_end_row + 1):
        quarter_cell = value_ws.cell(row_number, CORE_COLUMNS["quarter"].column)
        if _is_empty(quarter_cell.value):
            continue

        quarter_label = _format_quarter_label(quarter_cell.value)
        notes: list[str] = []
        additional_metrics: list[AdditionalMetricValue] = []

        shares_or_note = _clean_value(value_ws.cell(row_number, CORE_COLUMNS["shares_or_notes"].column).value)
        shares_millions = _to_float(shares_or_note)
        if shares_or_note is not None and shares_millions is None:
            notes.append(str(shares_or_note).strip())

        for column, metric_name in additional_headers.items():
            cell = value_ws.cell(row_number, column)
            value = _clean_value(cell.value)
            if value is None:
                continue

            normalized_name = metric_name or f"Additional metric {get_column_letter(column)}"
            if isinstance(value, str) and (metric_name is None or "note" in normalized_name.lower()):
                notes.append(value.strip())
                continue

            additional_metrics.append(
                AdditionalMetricValue(
                    name=normalized_name,
                    value=value,
                    column_letter=get_column_letter(column),
                    source_cell=cell.coordinate,
                )
            )

        _validate_formula_cells(formula_ws, row_number, warnings)
        quarter_date = _extract_quarter_date(quarter_cell.value)
        is_forecast = _is_forecast_row(
            formula_ws=formula_ws,
            row_number=row_number,
            quarter_label=quarter_label,
            quarter_date=quarter_date,
        )

        quarters.append(
            QuarterRecord(
                row_number=row_number,
                quarter_label=quarter_label,
                quarter_date=quarter_date,
                status=_to_text(value_ws.cell(row_number, CORE_COLUMNS["status"].column).value),
                revenue_estimate_millions=_to_float(value_ws.cell(row_number, CORE_COLUMNS["revenue_estimate"].column).value),
                revenue_actual_millions=_to_float(value_ws.cell(row_number, CORE_COLUMNS["revenue_actual"].column).value),
                rolling_4q_revenue_millions=_forecast_zero_as_missing(
                    _to_float(value_ws.cell(row_number, CORE_COLUMNS["rolling_revenue"].column).value),
                    is_forecast,
                ),
                rolling_4q_revenue_yoy_growth=_forecast_zero_as_missing(
                    _to_float(value_ws.cell(row_number, CORE_COLUMNS["rolling_revenue_yoy"].column).value),
                    is_forecast,
                ),
                workbook_revenue_surprise_pct=_to_float(value_ws.cell(row_number, CORE_COLUMNS["revenue_surprise"].column).value),
                workbook_revenue_qoq_growth=_to_float(value_ws.cell(row_number, CORE_COLUMNS["revenue_qoq"].column).value),
                workbook_revenue_yoy_growth=_to_float(value_ws.cell(row_number, CORE_COLUMNS["revenue_yoy"].column).value),
                eps_estimate=_to_float(value_ws.cell(row_number, CORE_COLUMNS["eps_estimate"].column).value),
                eps_actual=_to_float(value_ws.cell(row_number, CORE_COLUMNS["eps_actual"].column).value),
                rolling_4q_eps=_forecast_zero_as_missing(
                    _to_float(value_ws.cell(row_number, CORE_COLUMNS["rolling_eps"].column).value),
                    is_forecast,
                ),
                rolling_4q_eps_yoy_growth=_forecast_zero_as_missing(
                    _to_float(value_ws.cell(row_number, CORE_COLUMNS["rolling_eps_yoy"].column).value),
                    is_forecast,
                ),
                workbook_eps_surprise_pct=_to_float(value_ws.cell(row_number, CORE_COLUMNS["eps_surprise"].column).value),
                workbook_eps_qoq_growth=_to_float(value_ws.cell(row_number, CORE_COLUMNS["eps_qoq"].column).value),
                workbook_eps_yoy_growth=_to_float(value_ws.cell(row_number, CORE_COLUMNS["eps_yoy"].column).value),
                shares_millions=shares_millions,
                total_earnings_millions=_to_float(value_ws.cell(row_number, CORE_COLUMNS["total_earnings"].column).value),
                price_before_earnings=_to_float(value_ws.cell(row_number, CORE_COLUMNS["price_before"].column).value),
                price_after_earnings=_to_float(value_ws.cell(row_number, CORE_COLUMNS["price_after"].column).value),
                workbook_price_reaction_pct=_to_float(value_ws.cell(row_number, CORE_COLUMNS["price_reaction"].column).value),
                is_forecast=is_forecast,
                notes=tuple(notes),
                additional_metrics=tuple(additional_metrics),
            )
        )

    ignored_sheets = tuple(sheet for sheet in EXPECTED_SHEETS if sheet != DATA_SHEET_NAME)
    ignored_ranges = (
        f"{DATA_SHEET_NAME}!A2:AI2",
        f"{DATA_SHEET_NAME}!A5:AI5",
        f"{DATA_SHEET_NAME}!A{data_end_row + 1}:AI{value_ws.max_row}",
    )

    return ParsedWorkbook(
        company=company,
        quarters=tuple(quarters),
        ignored_sheets=ignored_sheets,
        ignored_ranges=ignored_ranges,
        warnings=tuple(warnings),
    )


def _validate_sheets(sheet_names: list[str]) -> None:
    missing = [sheet for sheet in EXPECTED_SHEETS if sheet not in sheet_names]
    if missing:
        raise WorkbookValidationError(
            f"Workbook is missing required sheet(s): {', '.join(missing)}. "
            f"Found sheets: {', '.join(sheet_names)}."
        )


def _find_header_row(ws: Worksheet) -> int:
    for row in HEADER_ROW_CANDIDATES:
        value = _to_text(ws.cell(row, CORE_COLUMNS["quarter"].column).value)
        if value and value.strip().lower() == "quarter":
            return row
    candidates = ", ".join(f"A{row}" for row in HEADER_ROW_CANDIDATES)
    raise WorkbookValidationError(
        f"Could not find the fixed quarterly table header. Expected 'Quarter' in {candidates}."
    )


def _validate_core_headers(ws: Worksheet, header_row: int) -> None:
    errors: list[str] = []
    for key, spec in CORE_COLUMNS.items():
        if key in {"status", "total_earnings", "shares_or_notes"}:
            continue
        header = _to_text(ws.cell(header_row, spec.column).value)
        normalized = _normalize_header(header)
        if not header:
            errors.append(f"{get_column_letter(spec.column)}{header_row} is blank; expected {spec.name!r}")
            continue
        for token in spec.required_header_tokens:
            if token not in normalized:
                errors.append(
                    f"{get_column_letter(spec.column)}{header_row} has {header!r}; "
                    f"expected a header containing {token!r}."
                )
                break

    if errors:
        raise WorkbookValidationError("Workbook headers do not match the expected template: " + " ".join(errors))


def _find_data_end_row(ws: Worksheet) -> int:
    last_data_row = DATA_START_ROW - 1
    for row_number in range(DATA_START_ROW, ws.max_row + 1):
        value = _to_text(ws.cell(row_number, CORE_COLUMNS["quarter"].column).value)
        if value and value.strip().lower().startswith(FOOTER_PREFIX):
            break
        if value:
            last_data_row = row_number
    return last_data_row


def _read_additional_headers(ws: Worksheet, header_row: int) -> dict[int, str | None]:
    headers: dict[int, str | None] = {}
    for column in range(22, ws.max_column + 1):
        header = _to_text(ws.cell(header_row, column).value)
        fallback_header = _to_text(ws.cell(header_row - 1, column).value) if header_row > 1 else None
        headers[column] = header or fallback_header
    return headers


def _read_static_cells(ws: Worksheet, header_row: int) -> list[StaticCell]:
    cells = [
        StaticCell(DATA_SHEET_NAME, "A1", _clean_value(ws["A1"].value), "Company title"),
        StaticCell(DATA_SHEET_NAME, f"A{header_row}", _clean_value(ws.cell(header_row, 1).value), "Quarter table header"),
    ]
    for coordinate in ("Q3", "Q4", "U4"):
        value = _clean_value(ws[coordinate].value)
        if value is not None:
            cells.append(StaticCell(DATA_SHEET_NAME, coordinate, value, "Template static information"))
    return cells


def _read_company_info(
    workbook_path: Path,
    sheet_names: tuple[str, ...],
    header_row: int,
    data_end_row: int,
    static_cells: tuple[StaticCell, ...],
) -> CompanyInfo:
    title = str(static_cells[0].value or "").strip()
    ticker, company_name, description = _parse_title(title)
    return CompanyInfo(
        title=title,
        ticker=ticker,
        company_name=company_name,
        description=description,
        workbook_path=workbook_path,
        worksheet_names=sheet_names,
        data_sheet_name=DATA_SHEET_NAME,
        header_row=header_row,
        data_start_row=DATA_START_ROW,
        data_end_row=data_end_row,
        static_cells=static_cells,
    )


def _parse_title(title: str) -> tuple[str | None, str | None, str | None]:
    clean_title = " ".join(title.split())
    if not clean_title:
        return None, None, None
    ticker = clean_title.split(" ", 1)[0].strip(" -")
    company_name = None
    description = None
    if "-" in clean_title and "(" in clean_title and ")" in clean_title:
        after_dash = clean_title.split("-", 1)[1].strip()
        company_name = after_dash.split("(", 1)[0].strip() or None
        start = clean_title.find("(")
        end = clean_title.rfind(")")
        description = clean_title[start + 1 : end].strip() or None
    elif "(" in clean_title and ")" in clean_title:
        start = clean_title.find("(")
        end = clean_title.rfind(")")
        company_name = clean_title[start + 1 : end].strip() or None
        description = clean_title[end + 1 :].strip(" -") or None
    elif "-" in clean_title:
        after_dash = clean_title.split("-", 1)[1].strip()
        company_name = after_dash or None
    return ticker or None, company_name, description


def _validate_formula_cells(ws: Worksheet, row_number: int, warnings: list[ParserWarning]) -> None:
    for key in FORMULA_COLUMNS:
        column = CORE_COLUMNS[key].column
        cell = ws.cell(row_number, column)
        if _is_empty(cell.value):
            continue
        if not _is_formula(cell):
            warnings.append(
                ParserWarning(
                    message=f"Expected a formula in calculated column {get_column_letter(column)} for row {row_number}.",
                    sheet_name=DATA_SHEET_NAME,
                    cell=cell.coordinate,
                )
            )


def _is_formula(cell: Cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _is_forecast_row(
    formula_ws: Worksheet,
    row_number: int,
    quarter_label: str,
    quarter_date: date | None,
) -> bool:
    if _is_incomplete_date_label(quarter_label):
        return True
    if quarter_date is not None and quarter_date > date.today():
        return True
    if quarter_date is not None:
        return False
    return any(
        _is_red_input_cell(formula_ws.cell(row_number, column))
        for column in FORECAST_INPUT_COLUMNS
    )


def _is_red_input_cell(cell: Cell) -> bool:
    if cell.value is None or _is_formula(cell):
        return False
    color = cell.font.color
    if color is None:
        return False
    if color.type == "rgb":
        return bool(color.rgb and color.rgb.upper() in {"FFFF0000", "00FF0000"})
    if color.type == "indexed":
        return color.indexed == 3
    return False


def _is_incomplete_date_label(label: str | None) -> bool:
    text = str(label or "").strip()
    return bool(text.endswith("-") and len(text) >= 5 and text[:4].isdigit())


def _normalize_header(value: str | None) -> str:
    return " ".join((value or "").lower().replace("\n", " ").split())


def _format_quarter_label(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _extract_quarter_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if len(text) >= 10:
            try:
                return datetime.strptime(text[:10], "%Y-%m-%d").date()
            except ValueError:
                return None
    return None


def _to_text(value: Any) -> str | None:
    value = _clean_value(value)
    if value is None:
        return None
    return str(value).strip()


def _to_float(value: Any) -> float | None:
    value = _clean_value(value)
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text or text.startswith("#"):
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _forecast_zero_as_missing(value: float | None, is_forecast: bool) -> float | None:
    if is_forecast and value == 0:
        return None
    return value


def _clean_value(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.startswith("#"):
            return None
        return stripped
    return value


def _is_empty(value: Any) -> bool:
    return _clean_value(value) is None
